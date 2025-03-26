import json
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os

with open("data/products.json", "r") as f:
    products = json.load(f)

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Products"

headers = ["ID", "Title", "Category", "Price", "Image"]
sheet.append(headers)

for product in products:
    row = [product["id"], product["title"], product["category"], product["price"]]

    image_path = f"data/images/{product['id']}.jpg"
    if os.path.exists(image_path):
        img = Image(image_path)
        img.width, img.height = 100, 100
        sheet.append(row + [""])
        sheet.add_image(img, f"E{sheet.max_row}")
    else:
        sheet.append(row + ["Image not found"])

workbook.save("data/products.xlsx")
print("Products successfully exported to data/products.xlsx")
