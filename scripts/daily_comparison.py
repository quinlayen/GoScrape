import argparse
import sqlite3
import pandas as pd
import datetime
import json
import sys
import os
import random
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def get_db_connection(test_mode):
    db_dir = "db"
    if not os.path.exists(db_dir):
        os.makedirs(db_dir)
    if test_mode:
        db_path = os.path.join(db_dir, "test_products.db")
    else:
        db_path = os.path.join(db_dir, "products.db")
    return sqlite3.connect(db_path)

def create_products_table(conn):
    query = """
    CREATE TABLE IF NOT EXISTS products (
        id TEXT,
        title TEXT,
        category TEXT,
        price TEXT,
        scrape_date DATETIME DEFAULT CURRENT_TIMESTAMP,
        owned INTEGER,
        image TEXT
    );
    """
    conn.execute(query)
    conn.commit()

def load_data_from_json_if_empty(conn):
    df_db = pd.read_sql_query("SELECT * FROM products", conn)
    if df_db.empty:
        json_file = "data/products.json"
        if not os.path.exists(json_file):
            print("No product data found in the database or JSON file. Exiting.")
            sys.exit(1)
        with open(json_file, "r") as f:
            json_data = json.load(f)
        df_json = pd.DataFrame(json_data)
        if 'scrape_date' not in df_json.columns:
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            df_json["scrape_date"] = current_time
        df_json.to_sql("products", conn, if_exists="append", index=False)
        df_db = pd.read_sql_query("SELECT * FROM products", conn)
    return df_db

def simulate_second_date_if_needed(conn, df_db, test_mode):
    unique_dates = df_db['scrape_date'].unique()
    if test_mode and len(unique_dates) < 2:
        base_date = unique_dates[0]
        try:
            base_dt = datetime.datetime.strptime(base_date, "%Y-%m-%d %H:%M:%S")
        except Exception as e:
            print(f"Error parsing base date: {e}")
            sys.exit(1)
        new_date = (base_dt - datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
        df_simulated = df_db.copy()
        df_simulated['scrape_date'] = new_date
        df_simulated['price'] = df_simulated['price'].apply(
            lambda x: f"${float(x.replace('$','')) * random.uniform(0.9, 1.1):.2f}"
        )
        df_simulated.to_sql("products", conn, if_exists="append", index=False)
        df_db = pd.read_sql_query("SELECT * FROM products", conn)
    return df_db

def get_comparison_data(conn):
    query = "SELECT DISTINCT scrape_date FROM products ORDER BY scrape_date DESC LIMIT 2;"
    dates = [row[0] for row in conn.execute(query)]
    if len(dates) < 2:
        print("Not enough data to perform a comparison.")
        sys.exit(1)
    date1, date2 = dates[0], dates[1]
    df_latest = pd.read_sql_query(f"SELECT * FROM products WHERE scrape_date = '{date1}'", conn)
    df_previous = pd.read_sql_query(f"SELECT * FROM products WHERE scrape_date = '{date2}'", conn)
    return df_latest, df_previous, date1, date2

def merge_comparison_data(df_latest, df_previous):
    comparison = pd.merge(df_latest, df_previous, on="id", suffixes=("_latest", "_previous"))
    def clean_price(price):
        try:
            return float(str(price).replace("$", "").replace(",", ""))
        except Exception:
            return None
    comparison["price_latest"] = comparison["price_latest"].apply(clean_price)
    comparison["price_previous"] = comparison["price_previous"].apply(clean_price)
    comparison["price_change"] = comparison["price_latest"] - comparison["price_previous"]
    return comparison

def prepare_comparison_df(comparison):
    final_columns = [
        "id", "title_latest", "category_latest", "owned_latest",
        "price_latest", "scrape_date_latest", "image_latest",
        "price_previous", "scrape_date_previous", "price_change"
    ]
    df_final = comparison[final_columns].copy()
    df_final.rename(columns={
        "title_latest": "title",
        "category_latest": "category",
        "owned_latest": "owned",
        "image_latest": "image"
    }, inplace=True)
    return df_final

def auto_adjust_dimensions(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].row == 1:
            ws.row_dimensions[row[0].row].height = 15
        else:
            ws.row_dimensions[row[0].row].height = 75
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        ws.column_dimensions[col_letter].width = max_length + 2

def write_excel_report(df_all, comparison):
    comp_df = prepare_comparison_df(comparison)
    with pd.ExcelWriter("daily_price_comparison.xlsx", engine='openpyxl') as writer:
        df_all.to_excel(writer, sheet_name="All Items", index=False)
        comp_df.to_excel(writer, sheet_name="Price Changes", index=False)
    wb = openpyxl.load_workbook("daily_price_comparison.xlsx")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    ws_all = wb["All Items"]
    header_all = {cell.value: cell.column for cell in ws_all[1]}
    img_col_all = header_all.get("image")
    if img_col_all is None:
        img_col_all = ws_all.max_column + 1
        ws_all.cell(row=1, column=img_col_all, value="Product Image")
    for row in ws_all.iter_rows(min_row=2, max_col=ws_all.max_column):
        image_filename = row[header_all["image"] - 1].value
        if image_filename:
            image_path = os.path.join("data", "images", image_filename)
            if os.path.exists(image_path):
                img = OpenpyxlImage(image_path)
                img.width = 100
                img.height = 100
                cell = ws_all.cell(row=row[0].row, column=img_col_all)
                ws_all.add_image(img, cell.coordinate)
    
    ws_changes = wb["Price Changes"]
    header_changes = {cell.value: cell.column for cell in ws_changes[1]}
    img_col_changes = header_changes.get("image")
    if img_col_changes is None:
        img_col_changes = ws_changes.max_column + 1
        ws_changes.cell(row=1, column=img_col_changes, value="Product Image")
    for row in ws_changes.iter_rows(min_row=2, max_col=ws_changes.max_column):
        image_filename = row[header_changes["image"] - 1].value
        if image_filename:
            image_path = os.path.join("data", "images", image_filename)
            if os.path.exists(image_path):
                img = OpenpyxlImage(image_path)
                img.width = 100
                img.height = 100
                cell = ws_changes.cell(row=row[0].row, column=img_col_changes)
                ws_changes.add_image(img, cell.coordinate)
        try:
            price_change_val = float(row[header_changes["price_change"] - 1].value)
        except:
            price_change_val = 0
        fill = red_fill if price_change_val > 0 else green_fill if price_change_val < 0 else None
        if fill:
            for cell in row:
                cell.fill = fill
    
    for ws in [ws_all, ws_changes]:
        auto_adjust_dimensions(ws)
    wb.save("daily_price_comparison.xlsx")

def main():
    parser = argparse.ArgumentParser(description="Daily price comparison")
    parser.add_argument("--test", "-t", action="store_true", help="Enable test mode")
    args = parser.parse_args()
    conn = get_db_connection(args.test)
    create_products_table(conn)
    df_db = load_data_from_json_if_empty(conn)
    df_db = simulate_second_date_if_needed(conn, df_db, args.test)
    df_latest, df_previous, date1, date2 = get_comparison_data(conn)
    comparison = merge_comparison_data(df_latest, df_previous)
    write_excel_report(df_db, comparison)
    print(f"Comparing dates: {date1} (latest) and {date2} (previous)")
    print("Comparison complete. Results saved to daily_price_comparison.xlsx")
    conn.close()

if __name__ == "__main__":
    main()